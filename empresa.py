import openpyxl

# Crie um novo arquivo Excel
arquivo_excel = openpyxl.Workbook()

# Selecione a planilha padrão (aba ativa)
planilha = arquivo_excel.active

# Cabeçalhos das colunas
cabecalhos = ['Funcionário', 'Contratação', 'Idade', 'Cargo', 'Setor']

# Adicione cabeçalhos para as colunas
for col_idx, cabecalho in enumerate(cabecalhos, start=1):
    planilha.cell(row=1, column=col_idx, value=cabecalho)

# Pergunte ao usuário quantas empresas deseja adicionar
num_empresas = int(input("Quantas empresas você deseja adicionar? "))

# Lista para armazenar as informações das empresas
empresas = []

# Preencha a lista com as informações das empresas
for i in range(num_empresas):
    nome = input(f"Nome da empresa {i + 1}: ")
    contratacao = input(f"Data de contratação da empresa {i + 1}: ")
    idade = input(f"Idade da empresa {i + 1}: ")
    cargo = input(f"Cargo da empresa {i + 1}: ")
    setor = input(f"Setor da empresa {i + 1}: ")

    empresa_info = [nome, contratacao, idade, cargo, setor]
    empresas.append(empresa_info)

# Preencha a planilha com as informações das empresas
for row_idx, empresa in enumerate(empresas, start=2):
    for col_idx, campo in enumerate(empresa, start=1):
        planilha.cell(row=row_idx, column=col_idx, value=campo)

# Solicite o nome do arquivo para salvar a planilha
nome_arquivo = input("Digite o nome do arquivo para salvar a planilha (ex: empresas.xlsx): ")

# Adicione a extensão .xlsx se não estiver presente no nome do arquivo
if not nome_arquivo.endswith(".xlsx"):
    nome_arquivo += ".xlsx"

# Salve a planilha em um arquivo
arquivo_excel.save(nome_arquivo)

# Feche o arquivo Excel
arquivo_excel.close()

print(f"A planilha foi salva como '{nome_arquivo}'.")
